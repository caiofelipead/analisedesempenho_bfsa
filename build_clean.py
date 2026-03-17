import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

with open('/tmp/team_data.json') as f:
    team = json.load(f)

wb = Workbook()

# ── STYLES (minimal) ──
hdr_fill = PatternFill('solid', fgColor='C9A227')
hdr_font = Font(name='Arial', bold=True, size=10, color='000000')
alt_fill = PatternFill('solid', fgColor='F2F2F2')
title_font = Font(name='Arial', bold=True, size=13, color='1A1A1A')
sub_font = Font(name='Arial', size=9, color='666666')
body_font = Font(name='Arial', size=10, color='1A1A1A')
bold_font = Font(name='Arial', bold=True, size=10, color='1A1A1A')
link_font = Font(name='Arial', size=10, color='0563C1', underline='single')
tb = Border(left=Side(style='thin', color='D0D0D0'), right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'), bottom=Side(style='thin', color='D0D0D0'))
center = Alignment(horizontal='center', vertical='center')
left = Alignment(horizontal='left', vertical='center')
wrap_left = Alignment(horizontal='left', vertical='center', wrap_text=True)

def hdr(ws, row, cs, headers, widths=None):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=cs+i, value=h)
        c.fill = hdr_fill; c.font = hdr_font; c.alignment = center; c.border = tb
    if widths:
        for i, w in enumerate(widths):
            ws.column_dimensions[get_column_letter(cs+i)].width = w

def dc(ws, row, col, val, alt=False):
    c = ws.cell(row=row, column=col, value=val)
    if alt: c.fill = alt_fill
    c.font = body_font; c.alignment = center; c.border = tb
    return c

squad = [
    (1,'Victor Souza','GK','Goleiro'),(2,'Jonathan Lemos','RB','Lateral Direito'),
    (3,'Éricson','CB','Zagueiro'),(4,'Gustavo Vilar','CB','Zagueiro'),
    (5,'Leandro Maciel','CDM','Volante'),(6,'Patrick Brey','LB','Lateral Esquerdo'),
    (7,'Kelvin Giacobe','RW','Extremo'),(8,'Éverton Morelli','CDM','Volante'),
    (9,'Hygor Cléber','ST','Atacante'),(10,'Rafael Gava','CAM','Meia'),
    (11,'Jéfferson Nem','LW','Extremo'),(12,'Jordan Esteves','GK','Goleiro'),
    (13,'Wallace Fortuna','CB','Zagueiro'),(14,'Carlos Eduardo (Carlão)','CB','Zagueiro'),
    (15,'Guilherme Mariano','CB','Zagueiro'),(16,'Matheus Sales','CDM','Volante'),
    (17,'Guilherme Queiróz','ST','Atacante'),(19,'Márcio Dutra (Maranhão)','RW','Extremo'),
    (20,'Marquinho','CAM','Meia'),(21,'Luizão','ST','Atacante'),
    (22,'Gabriel Inocêncio','RB','Lateral Direito'),(23,'Wesley Pinheiro','LW','Extremo'),
    (25,'Brenno Klippel','GK','Goleiro'),(26,'Felipe Vieira','LB','Lateral Esquerdo'),
    (27,'Darlan Batista','CB','Zagueiro'),(29,'Thiaguinho','CDM','Volante'),
    (30,'Zé Hugo','RW','Extremo'),(0,'Pedro Tortello','CDM','Volante'),
    (0,'Thalles','ST','Atacante'),(0,'Hebert Badaró','CB','Zagueiro'),
    (0,'Érik','CDM','Volante'),(0,'Adriano','GK','Goleiro'),
    (0,'Whalacy Ermeliano','LW','Extremo'),(0,'Yuri Felipe','CDM','Volante'),
    (0,'Henrique Teles','LB','Lateral Esquerdo'),(0,'Felipe Penha','CAM','Meia'),
    (0,'Pedrinho','RB','Lateral Direito'),
]

# ════════════════════════════════════════════════════════════
# 1. CADASTRO ATLETAS
# ════════════════════════════════════════════════════════════
ws = wb.active
ws.title = 'Cadastro Atletas'
ws.sheet_properties.tabColor = 'C9A227'

ws.merge_cells('B2:Z2')
ws.cell(row=2, column=2, value='CADASTRO DE ATLETAS — BOTAFOGO SP 2026').font = title_font
ws.merge_cells('B3:Z3')
ws.cell(row=3, column=2, value='Preencher dados básicos + colar links de foto e vídeos').font = sub_font

cols = ['Nº','Nome','Apelido','Posição','Grupo','Pé','Nasc.','Idade','Altura','Peso',
        'Nacionalidade','Contrato Até','Valor TM',
        'Link Foto','Link Vídeo Highlights','Link Vídeo Wyscout','Link Transfermarkt',
        'Status']
widths = [5,26,14,6,14,10,12,6,7,6,12,12,10,
          34,34,34,34,12]
hdr(ws, 5, 2, cols, widths)

dv_pos = DataValidation(type='list', formula1='"GK,RCB,CB,LCB,RB,LB,RWB,LWB,CDM,CM,CAM,RM,LM,RW,LW,CF,ST"')
dv_pe = DataValidation(type='list', formula1='"Direito,Esquerdo,Ambidestro"')
dv_st = DataValidation(type='list', formula1='"Disponível,Lesionado,Suspenso,Emprestado,Negociando,Sub-20"')
ws.add_data_validation(dv_pos); ws.add_data_validation(dv_pe); ws.add_data_validation(dv_st)

for i, (num, name, pos, pos_br) in enumerate(squad):
    r = 6 + i; alt = i % 2 == 1
    short = name.split('(')[-1].replace(')','').strip() if '(' in name else name.split()[-1] if len(name.split()) > 1 else name
    dc(ws, r, 2, num if num > 0 else '', alt)
    c = dc(ws, r, 3, name, alt); c.font = bold_font; c.alignment = left
    dc(ws, r, 4, short, alt)
    dc(ws, r, 5, pos, alt); dv_pos.add(ws.cell(row=r, column=5))
    dc(ws, r, 6, pos_br, alt)
    dc(ws, r, 7, '', alt); dv_pe.add(ws.cell(row=r, column=7))
    for j in range(7, 13): dc(ws, r, 2+j, '', alt)
    # Link columns
    for j in range(15, 19):
        c = dc(ws, r, j, '', alt); c.font = link_font; c.alignment = left
    dc(ws, r, 19, 'Disponível', alt); dv_st.add(ws.cell(row=r, column=19))

# Extra blank rows for new signings
for i in range(8):
    r = 6 + len(squad) + i; alt = (len(squad)+i) % 2 == 1
    for j in range(len(cols)): dc(ws, r, 2+j, '', alt)
    dv_pos.add(ws.cell(row=r, column=5)); dv_pe.add(ws.cell(row=r, column=7)); dv_st.add(ws.cell(row=r, column=19))

ws.freeze_panes = 'D6'
ws.column_dimensions['A'].width = 2
ws.auto_filter.ref = f'B5:S{6+len(squad)+7}'

# ════════════════════════════════════════════════════════════
# 2. DESEMPENHO COLETIVO
# ════════════════════════════════════════════════════════════
ws2 = wb.create_sheet('Coletivo')
ws2.sheet_properties.tabColor = '58A6FF'

ws2.merge_cells('B2:AH2')
ws2.cell(row=2, column=2, value='DESEMPENHO COLETIVO POR JOGO').font = title_font
ws2.merge_cells('B3:AH3')
ws2.cell(row=3, column=2, value='1 linha = 1 jogo | Wyscout Team Stats export → colar aqui').font = sub_font

cc = ['Comp','Rodada','Data','Adversário','Local','Res','Placar','Sistema',
      'GM','GS','xG','xGA','Posse%',
      'Remates','Rem Alvo','Rem%','Passes','Pass Crt','Pass%',
      'Cruzamentos','Cruz Crt','Cruz%','PPDA','Duelos%',
      'Intercep','Recup','Perdas','Ataq Pos','Contra-Ataq',
      'BP Rem','Toques Área','Intensidade','Faltas','Cart Am','Cart Vm']
cw = [10,7,12,26,5,5,8,16,4,4,5,5,7,7,7,6,7,7,6,8,7,6,6,7,7,7,7,7,8,6,9,9,6,6,6]
hdr(ws2, 5, 2, cc, cw)

dv_comp = DataValidation(type='list', formula1='"Paulistão,Série B,Amistoso"')
dv_loc = DataValidation(type='list', formula1='"C,F"')
dv_res = DataValidation(type='list', formula1='"V,E,D"')
ws2.add_data_validation(dv_comp); ws2.add_data_validation(dv_loc); ws2.add_data_validation(dv_res)

# Pre-fill Paulistão
for i, m in enumerate(team):
    r = 6+i; alt = i%2==1
    jogo = m['jogo']
    if jogo.startswith('Botafogo'):
        adv = jogo.split(' - ')[1].rsplit(' ',1)[0]; loc = 'C'
    else:
        adv = jogo.split(' - ')[0].strip(); loc = 'F'
    placar = jogo.split(' ')[-1] if ':' in jogo else ''
    vals = ['Paulistão',f'R{i+1}',m['data'],adv,loc,m['resultado'],placar,
            m['sistema'],m['gols_pro'],m['gols_contra'],m['xg'],m['xg_contra'],m['posse'],
            m['remates'],m['remates_alvo'],m['remates_pct'],m['passes_total'],m['passes_certos'],m['passes_pct'],
            m['cruzamentos'],m['cruzamentos_certos'],m['cruzamentos_pct'],m['ppda'],m['duelos_pct'],
            m['interceptacoes'],m['recuperacoes'],m['perdas'],m['ataques_posicionais'],m['contra_ataques'],
            m['bp_remates'],m['toques_area'],m['intensidade'],m['faltas'],m['cartoes_am'],m['cartoes_vm']]
    for j,v in enumerate(vals): dc(ws2, r, 2+j, v, alt)

# Empty rows (38 Série B + 10 extra)
for i in range(48):
    r = 6+len(team)+i; alt = (len(team)+i)%2==1
    for j in range(len(cc)): dc(ws2, r, 2+j, '', alt)
    dv_comp.add(ws2.cell(row=r, column=2)); dv_loc.add(ws2.cell(row=r, column=6)); dv_res.add(ws2.cell(row=r, column=7))

ws2.freeze_panes = 'E6'
ws2.column_dimensions['A'].width = 2
ws2.auto_filter.ref = f'B5:AJ{6+len(team)+47}'

# ════════════════════════════════════════════════════════════
# 3. DESEMPENHO INDIVIDUAL — Wyscout Player Stats (73 colunas)
# ════════════════════════════════════════════════════════════
ws3 = wb.create_sheet('Individual')
ws3.sheet_properties.tabColor = '3FB950'

ws3.merge_cells('B2:BU2')
ws3.cell(row=2, column=2, value='DESEMPENHO INDIVIDUAL POR JOGO — WYSCOUT PLAYER STATS').font = title_font
ws3.merge_cells('B3:BU3')
ws3.cell(row=3, column=2, value='Copiar/colar direto do export Wyscout Player Stats | 73 colunas | Colunas pareadas: col N = total, col N+1 = sucesso').font = sub_font

# 73 colunas exatas do Wyscout Player Stats export
# Colunas pareadas: header na col N, col N+1 em branco (valor de "sucesso/certos")
ic_headers = [
    'Atleta', 'Jogo', 'Competition', 'Date', 'Posição', 'Minutos jogados',
    'Ações totais/bem sucedidos', '',  # cols 7-8
    'Golos', 'Assistências',
    'Remates', 'à baliza',  # cols 11-12
    'xG',
    'Passes', 'certos',  # cols 14-15
    'Passes longos', 'certos',  # cols 16-17
    'Cruzamentos', 'certos',  # cols 18-19
    'Dribbles', 'com sucesso',  # cols 20-21
    'Duelos', 'ganhos',  # cols 22-23
    'Duelos aéreos', 'ganhos',  # cols 24-25
    'Intercepções',
    'Perdas', 'no seu meio-campo',  # cols 27-28
    'Recuperações', 'no meio-campo do adversário',  # cols 29-30
    'Cartão amarelo', 'Cartão vermelho',
    'Duelos defensivos', 'ganhos',  # cols 33-34
    'Duelos de bola livre', 'won',  # cols 35-36
    'Carrinhos', 'bem sucedidos',  # cols 37-38
    'Alívios', 'Faltas', 'Cartões amarelos', 'Cartões vermelhos',
    'Assistências para remate',
    'Duelos ofensivos', 'ganhos',  # cols 44-45
    'Toques na área', 'Foras de jogo', 'Corridas seguidas', 'Faltas sofridas',
    'Passes em profundidade', 'certos',  # cols 50-51
    'xA', 'Segundas assistências',
    'Passes para terço final', 'certos',  # cols 54-55
    'Passes para a grande área', 'precisos',  # cols 56-57
    'Passes recebidos',
    'Passes para a frente', 'certos',  # cols 59-60
    'Passes para trás', 'certos',  # cols 61-62
    'Golos sofridos (GK)', 'xCG (GK)', 'Remates sofridos (GK)',
    'Defesas', 'com reflexos (GK)',  # cols 66-67
    'Saídas (GK)',
    'Carrinhos', 'bem sucedidos (GK)',  # cols 69-70
    'Pontapés de baliza (GK)', 'Pontapés de baliza curtos (GK)', 'Pontapés de baliza longos (GK)',
]

# Larguras para 73 colunas
iw = [
    20, 26, 12, 12, 10, 8,  # Atleta-Minutos
    10, 8,  # Ações totais/sucesso
    5, 7,   # Gols, Assists
    7, 7,   # Remates/baliza
    5,       # xG
    7, 7,   # Passes/certos
    8, 7,   # Passes longos/certos
    8, 7,   # Cruzamentos/certos
    7, 7,   # Dribles/sucesso
    7, 7,   # Duelos/ganhos
    8, 7,   # Duelos aéreos/ganhos
    7,       # Intercepções
    7, 9,   # Perdas/meio-campo
    8, 12,  # Recup/campo adv
    7, 7,   # Cartões
    9, 7,   # Duelos def/ganhos
    10, 7,  # Duelos bola livre/won
    8, 8,   # Carrinhos/sucesso
    7, 6, 7, 7,  # Alívios, Faltas, Cart Am, Cart Vm
    9,       # Assist remate
    9, 7,   # Duelos ofensivos/ganhos
    7, 7, 8, 7,  # Toques, Fora jogo, Corridas, Faltas sofridas
    10, 7,  # Passes profundidade/certos
    5, 8,   # xA, 2nd assists
    10, 7,  # Passes terço final/certos
    11, 7,  # Passes grande área/precisos
    8,       # Passes recebidos
    10, 7,  # Passes frente/certos
    9, 7,   # Passes trás/certos
    8, 6, 9,  # GK: gols sofridos, xCG, remates sofridos
    7, 8,   # GK: defesas/reflexos
    7,       # GK: saídas
    8, 9,   # GK: carrinhos/sucesso
    9, 10, 10,  # GK: pontapés baliza / curtos / longos
]

hdr(ws3, 5, 2, ic_headers, iw)

# 1000 rows for the season
for i in range(1000):
    r = 6+i; alt = i%2==1
    for j in range(len(ic_headers)): dc(ws3, r, 2+j, '', alt)

ws3.freeze_panes = 'C6'
ws3.column_dimensions['A'].width = 2
ws3.auto_filter.ref = f'B5:BU1005'

# ════════════════════════════════════════════════════════════
# 4. VÍDEOS
# ════════════════════════════════════════════════════════════
ws4 = wb.create_sheet('Vídeos')
ws4.sheet_properties.tabColor = 'F85149'

ws4.merge_cells('B2:L2')
ws4.cell(row=2, column=2, value='BIBLIOTECA DE VÍDEOS').font = title_font
ws4.merge_cells('B3:L3')
ws4.cell(row=3, column=2, value='Jogos completos, highlights, preleções, treinos').font = sub_font

vc = ['Tipo','Comp','Rodada','Data','Adversário/Descrição','Link Vídeo','Link Alternativo','Responsável','Notas']
vw = [14,10,7,12,28,36,36,12,28]
hdr(ws4, 5, 2, vc, vw)

dv_tipo = DataValidation(type='list', formula1='"Jogo Completo,Highlights,Preleção,Análise Tática,Bolas Paradas,Treino,Individual,Material Orientador"')
ws4.add_data_validation(dv_tipo)

for i in range(100):
    r = 6+i; alt = i%2==1
    dc(ws4, r, 2, '', alt); dv_tipo.add(ws4.cell(row=r, column=2))
    for j in range(1, len(vc)): dc(ws4, r, 2+j, '', alt)
    # Link columns with link font
    ws4.cell(row=r, column=7).font = link_font; ws4.cell(row=r, column=7).alignment = left
    ws4.cell(row=r, column=8).font = link_font; ws4.cell(row=r, column=8).alignment = left

ws4.freeze_panes = 'F6'
ws4.column_dimensions['A'].width = 2
ws4.auto_filter.ref = f'B5:J105'

# ════════════════════════════════════════════════════════════
# 5. CALENDÁRIO
# ════════════════════════════════════════════════════════════
ws5 = wb.create_sheet('Calendário')
ws5.sheet_properties.tabColor = 'DB6D28'

ws5.merge_cells('B2:O2')
ws5.cell(row=2, column=2, value='CALENDÁRIO OPERACIONAL 2026').font = title_font
ws5.merge_cells('B3:O3')
ws5.cell(row=3, column=2, value='Paulistão + Série B | Controle de processos por rodada').font = sub_font

ch = ['Comp','Rodada','Data','Adversário','Local','ADV','PRE','POS','DAT','WYS','TRE','BSP','IND','Obs']
cwi = [10,7,12,26,5,5,5,5,5,5,5,5,5,20]
hdr(ws5, 5, 2, ch, cwi)

dv_proc = DataValidation(type='list', formula1='"✓,⏳,—"', allow_blank=True)
ws5.add_data_validation(dv_proc)

paul = [('Paulistão','R1','11/01','Velo Clube','F'),('Paulistão','R2','15/01','Noroeste','C'),
        ('Paulistão','R3','18/01','Red Bull Bragantino','F'),('Paulistão','R4','23/01','Primavera SP','C'),
        ('Paulistão','R5','25/01','Novorizontino','F'),('Paulistão','R6','02/02','Palmeiras','C'),
        ('Paulistão','R7','07/02','Guarani','F'),('Paulistão','R8','16/02','Capivariano','C')]

sb1 = [('Série B','R1','21/03','Fortaleza','C'),('Série B','R2','01/04','América-MG','F'),
       ('Série B','R3','05/04','São Bernardo','C'),('Série B','R4','10/04','Criciúma','F'),
       ('Série B','R5','19/04','Atlético-GO','C'),('Série B','R6','26/04','Cuiabá','F'),
       ('Série B','R7','03/05','Náutico','C'),('Série B','R8','10/05','Novorizontino','F'),
       ('Série B','R9','17/05','Goiás','F'),('Série B','R10','24/05','Athletic','C'),
       ('Série B','R11','31/05','Ponte Preta','F'),('Série B','R12','10/06','Vila Nova-GO','F'),
       ('Série B','R13','14/06','Operário-PR','C'),('Série B','R14','21/06','Ceará','F'),
       ('Série B','R15','28/06','CRB','C'),('Série B','R16','05/07','Avaí','C'),
       ('Série B','R17','12/07','Sport','F'),('Série B','R18','19/07','Londrina','F'),
       ('Série B','R19','26/07','Juventude','C')]

sb2_dates = ['02/08','09/08','16/08','23/08','30/08','06/09','13/09','20/09','27/09',
             '04/10','11/10','18/10','25/10','01/11','08/11','15/11','18/11','19/11','20/11']

all_games = paul[:]
for g in sb1: all_games.append(g)
for i,(_,_,adv,loc) in enumerate([(c,r,a,l) for c,r,a,l in [(g[0],g[1],g[3],g[4]) for g in sb1]]):
    inv = 'F' if loc=='C' else 'C'
    all_games.append(('Série B', f'R{20+i}', sb2_dates[i], adv, inv))

for i, (comp, rod, dt, adv, loc) in enumerate(all_games):
    r = 6+i; alt = i%2==1
    dc(ws5, r, 2, comp, alt)
    dc(ws5, r, 3, rod, alt)
    dc(ws5, r, 4, dt, alt)
    c = dc(ws5, r, 5, adv, alt); c.font = bold_font
    dc(ws5, r, 6, loc, alt)
    done = comp == 'Paulistão'
    for j in range(7, 15):
        c = dc(ws5, r, j, '✓' if done else '', alt)
        if not done: dv_proc.add(c)
    dc(ws5, r, 15, '', alt)

ws5.freeze_panes = 'E6'
ws5.column_dimensions['A'].width = 2
ws5.auto_filter.ref = f'B5:O{6+len(all_games)-1}'

# ════════════════════════════════════════════════════════════
# SAVE
# ════════════════════════════════════════════════════════════
output = '/sessions/relaxed-peaceful-faraday/mnt/outputs/BFSA_Dados_2026.xlsx'
wb.save(output)
print(f'Saved: {output}')
print(f'Sheets: {wb.sheetnames}')
print(f'Total: {len(wb.sheetnames)} abas')
